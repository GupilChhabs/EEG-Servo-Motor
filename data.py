import os
import openpyxl
path="C:/Users/gupil/OneDrive/Desktop/close"
ent=os.listdir(path)
print(ent)
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(len(ent)):
	c1 = sheet.cell(row = i+1, column = 1)
	c1.value = ent[i]

path="C:/Users/gupil/OneDrive/Desktop/yes"
ent=os.listdir(path)
print(ent)
for i in range(len(ent)):
	c2 = sheet.cell(row = i+1, column = 2)
	c2.value = ent[i]
wb.save(filename='EEG.xlsx')
